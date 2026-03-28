"""Executive report router — generate HTML and JSON summary reports."""

from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from netlanventory.api.dependencies import get_current_active_user, get_db
from netlanventory.core.logging import get_logger
from netlanventory.models.asset import Asset
from netlanventory.models.asset_cve import AssetCve
from netlanventory.models.cve import Cve
from netlanventory.models.default_creds_report import DefaultCredsReport
from netlanventory.models.nuclei_report import NucleiReport
from netlanventory.models.ssh_audit_report import SshAuditReport
from netlanventory.models.ssl_scan_report import SslScanReport
from netlanventory.models.testssl_report import TestsslReport
from netlanventory.models.zap_report import ZapReport

logger = get_logger(__name__)

router = APIRouter(prefix="/reports", tags=["reports"])

DbDep = Annotated[AsyncSession, Depends(get_db)]
AuthDep = Annotated[object, Depends(get_current_active_user)]


# ── Data aggregation ──────────────────────────────────────────────────────────


async def _collect_report_data(db: AsyncSession) -> dict:
    """Collect all data needed for the executive report."""
    now = datetime.now(timezone.utc)
    thirty_days_ago = now - timedelta(days=30)

    # Total assets
    total_assets = (
        await db.execute(select(func.count()).select_from(Asset))
    ).scalar_one()
    active_assets = (
        await db.execute(
            select(func.count()).select_from(Asset).where(Asset.is_active.is_(True))
        )
    ).scalar_one()

    # Assets by criticality
    crit_result = await db.execute(
        select(Asset.criticality, func.count().label("cnt"))
        .group_by(Asset.criticality)
    )
    assets_by_criticality = {row[0]: row[1] for row in crit_result.all()}

    # CVE statistics
    total_cves = (
        await db.execute(select(func.count()).select_from(AssetCve))
    ).scalar_one()
    critical_cve_links = (
        await db.execute(
            select(func.count())
            .select_from(AssetCve)
            .join(Cve, AssetCve.cve_id == Cve.id)
            .where(Cve.severity == "Critical")
        )
    ).scalar_one()

    # Top 10 critical CVEs (by number of affected assets)
    top_cves_result = await db.execute(
        select(Cve.cve_id, Cve.severity, Cve.cvss_score, func.count(AssetCve.id).label("cnt"))
        .join(AssetCve, Cve.id == AssetCve.cve_id)
        .where(Cve.severity.in_(["Critical", "High"]))
        .group_by(Cve.cve_id, Cve.severity, Cve.cvss_score)
        .order_by(func.count(AssetCve.id).desc())
        .limit(10)
    )
    top_cves = [
        {"cve_id": r[0], "severity": r[1], "cvss_score": r[2], "affected_assets": r[3]}
        for r in top_cves_result.all()
    ]

    # SLA breaches
    sla_breaches = (
        await db.execute(
            select(func.count()).select_from(AssetCve).where(AssetCve.sla_breached.is_(True))
        )
    ).scalar_one()

    # CVEs discovered in last 30 days
    new_cves_30d = (
        await db.execute(
            select(func.count())
            .select_from(AssetCve)
            .where(AssetCve.discovered_at >= thirty_days_ago)
        )
    ).scalar_one()

    # Top risky assets (by risk_score)
    top_assets_result = await db.execute(
        select(Asset.id, Asset.name, Asset.ip, Asset.criticality, Asset.risk_score)
        .where(Asset.risk_score.isnot(None))
        .order_by(Asset.risk_score.desc())
        .limit(10)
    )
    top_assets = [
        {
            "id": str(r[0]),
            "name": r[1],
            "ip": r[2],
            "criticality": r[3],
            "risk_score": r[4],
        }
        for r in top_assets_result.all()
    ]

    return {
        "generated_at": now.isoformat(),
        "report_date": date.today().isoformat(),
        "assets": {
            "total": total_assets,
            "active": active_assets,
            "by_criticality": assets_by_criticality,
        },
        "cves": {
            "total_links": total_cves,
            "critical_count": critical_cve_links,
            "new_last_30d": new_cves_30d,
        },
        "sla": {
            "breaches": sla_breaches,
        },
        "top_cves": top_cves,
        "top_risky_assets": top_assets,
        "recommendations": _generate_recommendations(
            critical_cve_links=critical_cve_links,
            sla_breaches=sla_breaches,
            total_assets=total_assets,
        ),
    }


def _generate_recommendations(
    critical_cve_links: int,
    sla_breaches: int,
    total_assets: int,
) -> list[str]:
    recs = []
    if critical_cve_links > 0:
        recs.append(
            f"Immediately patch or mitigate {critical_cve_links} critical CVE(s) "
            "across your assets."
        )
    if sla_breaches > 0:
        recs.append(
            f"{sla_breaches} CVE remediation(s) have breached their SLA deadline. "
            "Review and act immediately."
        )
    if total_assets > 0:
        recs.append(
            "Ensure all assets have SSH or Nuclei scans scheduled for continuous vulnerability "
            "coverage."
        )
    recs.append(
        "Review asset criticality ratings and ensure risk scores are up to date."
    )
    recs.append(
        "Configure SLA deadlines and set up webhook notifications for critical CVEs."
    )
    return recs


# ── HTML generation ───────────────────────────────────────────────────────────


def _render_html(data: dict) -> str:
    """Generate a self-contained HTML executive report from data dict."""
    assets_by_crit = data["assets"]["by_criticality"]
    crit_rows = "".join(
        f"<tr><td>{k.capitalize()}</td><td>{v}</td></tr>"
        for k, v in sorted(assets_by_crit.items())
    )

    top_cve_rows = "".join(
        f"<tr>"
        f"<td>{c['cve_id']}</td>"
        f"<td><span class='sev sev-{(c['severity'] or '').lower()}'>{c['severity']}</span></td>"
        f"<td>{c['cvss_score'] if c['cvss_score'] is not None else 'N/A'}</td>"
        f"<td>{c['affected_assets']}</td>"
        f"</tr>"
        for c in data["top_cves"]
    )

    top_asset_rows = "".join(
        f"<tr>"
        f"<td>{a['name'] or 'N/A'}</td>"
        f"<td>{a['ip'] or 'N/A'}</td>"
        f"<td>{a['criticality'].capitalize()}</td>"
        f"<td>{a['risk_score']:.1f}</td>"
        f"</tr>"
        for a in data["top_risky_assets"]
    )

    rec_items = "".join(f"<li>{r}</li>" for r in data["recommendations"])

    sla_breach_color = "danger" if data["sla"]["breaches"] > 0 else "ok"
    crit_cve_color = "danger" if data["cves"]["critical_count"] > 0 else "ok"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NetLanVentory — Executive Report</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #f4f6f9; color: #333; }}
  header {{ background: #1a2332; color: #fff; padding: 24px 40px; }}
  header h1 {{ font-size: 1.8rem; }}
  header p {{ opacity: .7; font-size: 0.9rem; margin-top: 4px; }}
  main {{ max-width: 1100px; margin: 32px auto; padding: 0 24px; }}
  .kpis {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
           gap: 16px; margin-bottom: 32px; }}
  .kpi {{ background: #fff; border-radius: 8px; padding: 20px;
          box-shadow: 0 1px 4px rgba(0,0,0,.1); text-align: center; }}
  .kpi .value {{ font-size: 2.4rem; font-weight: 700; }}
  .kpi .label {{ font-size: 0.8rem; color: #666; margin-top: 4px; }}
  .kpi.danger .value {{ color: #d63031; }}
  .kpi.ok .value {{ color: #00b894; }}
  section {{ background: #fff; border-radius: 8px; padding: 24px;
             box-shadow: 0 1px 4px rgba(0,0,0,.1); margin-bottom: 24px; }}
  section h2 {{ font-size: 1.1rem; margin-bottom: 16px; color: #1a2332; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.9rem; }}
  th {{ background: #f0f2f5; text-align: left; padding: 8px 12px; font-weight: 600; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #eee; }}
  tr:last-child td {{ border-bottom: none; }}
  .sev {{ display: inline-block; padding: 2px 8px; border-radius: 4px;
          font-size: 0.75rem; font-weight: 600; text-transform: uppercase; }}
  .sev-critical {{ background: #ffeaea; color: #c0392b; }}
  .sev-high {{ background: #fff3e0; color: #e67e22; }}
  .sev-medium {{ background: #fffde7; color: #f39c12; }}
  .sev-low {{ background: #e8f5e9; color: #27ae60; }}
  ul.recs {{ padding-left: 20px; }}
  ul.recs li {{ margin-bottom: 8px; line-height: 1.5; }}
  footer {{ text-align: center; color: #999; font-size: 0.8rem; padding: 24px; }}
</style>
</head>
<body>
<header>
  <h1>NetLanVentory — Executive Security Report</h1>
  <p>Generated: {data['generated_at']} &nbsp;|&nbsp; Report date: {data['report_date']}</p>
</header>
<main>

<div class="kpis">
  <div class="kpi">
    <div class="value">{data['assets']['total']}</div>
    <div class="label">Total Assets</div>
  </div>
  <div class="kpi">
    <div class="value">{data['assets']['active']}</div>
    <div class="label">Active Assets</div>
  </div>
  <div class="kpi {crit_cve_color}">
    <div class="value">{data['cves']['critical_count']}</div>
    <div class="label">Critical CVE Links</div>
  </div>
  <div class="kpi">
    <div class="value">{data['cves']['total_links']}</div>
    <div class="label">Total CVE Links</div>
  </div>
  <div class="kpi {sla_breach_color}">
    <div class="value">{data['sla']['breaches']}</div>
    <div class="label">SLA Breaches</div>
  </div>
  <div class="kpi">
    <div class="value">{data['cves']['new_last_30d']}</div>
    <div class="label">New CVEs (30d)</div>
  </div>
</div>

<section>
  <h2>Assets by Criticality</h2>
  <table>
    <thead><tr><th>Criticality</th><th>Count</th></tr></thead>
    <tbody>{crit_rows}</tbody>
  </table>
</section>

<section>
  <h2>Top Critical / High CVEs</h2>
  <table>
    <thead>
      <tr><th>CVE ID</th><th>Severity</th><th>CVSS</th><th>Affected Assets</th></tr>
    </thead>
    <tbody>{top_cve_rows if top_cve_rows else '<tr><td colspan="4">No critical/high CVEs found.</td></tr>'}</tbody>
  </table>
</section>

<section>
  <h2>Top Risky Assets</h2>
  <table>
    <thead>
      <tr><th>Name</th><th>IP</th><th>Criticality</th><th>Risk Score</th></tr>
    </thead>
    <tbody>{top_asset_rows if top_asset_rows else '<tr><td colspan="4">No risk scores computed yet.</td></tr>'}</tbody>
  </table>
</section>

<section>
  <h2>Recommendations</h2>
  <ul class="recs">{rec_items}</ul>
</section>

</main>
<footer>NetLanVentory &mdash; Confidential &mdash; {data['report_date']}</footer>
</body>
</html>"""


# ── Endpoints ─────────────────────────────────────────────────────────────────


@router.get("/executive", response_class=HTMLResponse)
async def executive_report_html(db: DbDep, _auth: AuthDep) -> HTMLResponse:
    """Generate and return a full executive security report as HTML."""
    data = await _collect_report_data(db)
    html = _render_html(data)
    return HTMLResponse(content=html)


@router.get("/executive/data")
async def executive_report_data(db: DbDep, _auth: AuthDep) -> dict:
    """Return executive report data as JSON."""
    return await _collect_report_data(db)


# ── Per-asset report ──────────────────────────────────────────────────────────


@router.get("/assets/{asset_id}/data")
async def asset_report_data(asset_id: uuid.UUID, db: DbDep, _auth: AuthDep) -> dict:
    """Return a complete security audit report for one asset as JSON."""
    return await _collect_asset_report(db, asset_id)


@router.get("/assets/{asset_id}", response_class=HTMLResponse)
async def asset_report_html(
    asset_id: uuid.UUID, db: DbDep, _auth: AuthDep
) -> HTMLResponse:
    """Generate a complete per-asset security audit report as HTML."""
    data = await _collect_asset_report(db, asset_id)
    return HTMLResponse(content=_render_asset_html(data))


@router.get("/export/pdf")
async def export_pdf(db: DbDep, _auth: AuthDep):
    """Export the executive security report as PDF (requires WeasyPrint)."""
    try:
        from weasyprint import HTML as WeasyHTML
    except ImportError:
        raise HTTPException(status_code=501, detail="WeasyPrint not installed. Run: pip install netlanventory[reports]")

    from fastapi.responses import Response
    data = await _collect_report_data(db)
    html_content = _render_html(data)
    pdf_bytes = WeasyHTML(string=html_content).write_pdf()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=netlanventory_report.pdf"},
    )


@router.get("/export/pdf/executive")
async def export_pdf_executive(db: DbDep, _auth: AuthDep):
    """Export the RSSI executive summary as PDF."""
    try:
        from weasyprint import HTML as WeasyHTML
    except ImportError:
        raise HTTPException(status_code=501, detail="WeasyPrint not installed. Run: pip install netlanventory[reports]")

    from fastapi.responses import Response
    from netlanventory.api.routers.executive import get_executive_summary
    summary = await get_executive_summary(db)

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>body{{font-family:Arial,sans-serif;margin:40px;}}
h1{{color:#c0392b;}}table{{border-collapse:collapse;width:100%}}
td,th{{border:1px solid #ddd;padding:8px;}}th{{background:#2c3e50;color:white;}}</style>
</head><body>
<h1>NetLanVentory — Rapport Exécutif RSSI</h1>
<p>Généré le {summary.generated_at[:10]}</p>
<h2>Score de Risque Global: {summary.global_risk_score}/100 — Tendance: {summary.risk_trend}</h2>
<table><tr><th>Métrique</th><th>Valeur</th></tr>
<tr><td>Assets totaux</td><td>{summary.total_assets}</td></tr>
<tr><td>Assets actifs</td><td>{summary.active_assets}</td></tr>
<tr><td>CVEs totaux</td><td>{summary.total_cves}</td></tr>
<tr><td>CVEs critiques</td><td>{summary.critical_cves}</td></tr>
<tr><td>CVEs High</td><td>{summary.high_cves}</td></tr>
<tr><td>% non acquittés</td><td>{summary.unacknowledged_pct}%</td></tr>
<tr><td>Taux remédiation 30j</td><td>{summary.remediation_rate_pct}%</td></tr>
<tr><td>Couverture scan</td><td>{summary.coverage.scan_coverage_pct}%</td></tr>
<tr><td>Couverture hardening</td><td>{summary.coverage.hardening_coverage_pct}%</td></tr>
</table>
<h2>Top Assets à Risque</h2>
<table><tr><th>IP</th><th>Nom</th><th>Criticité</th><th>CVEs</th><th>CVEs Critiques</th></tr>
{"".join(f"<tr><td>{a.ip or ''}</td><td>{a.name or ''}</td><td>{a.criticality}</td><td>{a.cve_count}</td><td>{a.critical_cve_count}</td></tr>" for a in summary.top_risky_assets)}
</table>
</body></html>"""

    pdf_bytes = WeasyHTML(string=html).write_pdf()
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=executive_report.pdf"},
    )


@router.get("/export/xlsx")
async def export_xlsx(db: DbDep, _auth: AuthDep):
    """Export assets, CVEs and remediation as Excel spreadsheet."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed. Run: pip install netlanventory[reports]")

    import io
    from fastapi.responses import Response
    from sqlalchemy.orm import selectinload

    wb = openpyxl.Workbook()

    # Sheet 1: Assets
    ws_assets = wb.active
    ws_assets.title = "Assets"
    ws_assets.append(["ID", "IP", "Hostname", "Name", "Criticality", "OS", "Is Active", "Last Seen", "Risk Score"])
    asset_rows = (await db.execute(select(Asset).order_by(Asset.ip))).scalars().all()
    for a in asset_rows:
        ws_assets.append([
            str(a.id), a.ip or "", a.hostname or "", a.name or "",
            a.criticality or "", a.os_name or "", str(a.is_active),
            a.last_seen.isoformat() if a.last_seen else "",
            a.risk_score or "",
        ])

    # Sheet 2: CVEs
    ws_cves = wb.create_sheet("CVEs")
    ws_cves.append(["CVE ID", "Severity", "CVSS", "Asset IP", "Status", "Discovered At"])
    cve_rows = (await db.execute(
        select(AssetCve, Cve, Asset)
        .join(Cve, AssetCve.cve_id == Cve.id)
        .join(Asset, AssetCve.asset_id == Asset.id)
        .order_by(Cve.cvss_score.desc().nullslast())
        .limit(5000)
    )).all()
    for ac, cve, asset in cve_rows:
        ws_cves.append([
            cve.cve_id or "", cve.severity or "", cve.cvss_score or "",
            asset.ip or "", ac.ack_status or "",
            ac.discovered_at.isoformat() if ac.discovered_at else "",
        ])

    # Sheet 3: Remediation summary
    ws_rem = wb.create_sheet("Remediation")
    ws_rem.append(["CVE ID", "Severity", "CVSS", "Remediation", "Acknowledged Count", "Unacknowledged Count"])
    from sqlalchemy import case as sa_case
    rem_rows = (await db.execute(
        select(
            Cve.cve_id, Cve.severity, Cve.cvss_score, Cve.remediation,
            func.sum(sa_case((AssetCve.ack_status != "none", 1), else_=0)).label("acked"),
            func.sum(sa_case((AssetCve.ack_status == "none", 1), else_=0)).label("unacked"),
        )
        .join(AssetCve, AssetCve.cve_id == Cve.id)
        .group_by(Cve.cve_id, Cve.severity, Cve.cvss_score, Cve.remediation)
        .order_by(Cve.cvss_score.desc().nullslast())
        .limit(2000)
    )).all()
    for r in rem_rows:
        ws_rem.append([r.cve_id or "", r.severity or "", r.cvss_score or "", r.remediation or "", r.acked or 0, r.unacked or 0])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=netlanventory_export.xlsx"},
    )


# ── Asset report data aggregation ─────────────────────────────────────────────


async def _collect_asset_report(db: AsyncSession, asset_id: uuid.UUID) -> dict:
    """Aggregate all findings from all scan sources for one asset."""
    # Load asset with all relations
    result = await db.execute(
        select(Asset)
        .options(
            selectinload(Asset.ports),
            selectinload(Asset.cves).selectinload(AssetCve.cve),
            selectinload(Asset.tags),
        )
        .where(Asset.id == asset_id)
    )
    asset = result.scalar_one_or_none()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    now = datetime.now(timezone.utc)
    findings: list[dict] = []

    # ── CVE findings ──────────────────────────────────────────────────────────
    for link in (asset.cves or []):
        cve = link.cve
        if not cve:
            continue
        if link.ack_status in ("false_positive", "accepted"):
            continue

        priority = _cve_priority(cve, link)
        remediation = _cve_remediation(cve, link)

        findings.append({
            "id": str(link.id),
            "source": link.source or "unknown",
            "category": "cve",
            "severity": cve.severity or "Unknown",
            "title": cve.cve_id,
            "description": cve.description or f"Vulnerability {cve.cve_id}",
            "detail": {
                "cvss_score": cve.cvss_score,
                "epss_score": cve.epss_score,
                "epss_percentile": cve.epss_percentile,
                "kev": bool(cve.kev_date_added),
                "exploit_maturity": cve.exploit_maturity,
                "exploit_verified": link.exploit_verified,
                "package": link.package_name,
                "package_version": link.package_version,
                "fixed_version": link.fixed_version,
                "sla_deadline": link.sla_deadline.isoformat() if link.sla_deadline else None,
                "sla_breached": link.sla_breached,
            },
            "remediation": remediation,
            "priority": priority,
        })

    # ── Default credentials ───────────────────────────────────────────────────
    dc_result = await db.execute(
        select(DefaultCredsReport)
        .where(
            DefaultCredsReport.asset_id == asset_id,
            DefaultCredsReport.status == "completed",
        )
        .order_by(DefaultCredsReport.created_at.desc())
        .limit(1)
    )
    dc_report = dc_result.scalar_one_or_none()
    if dc_report and dc_report.findings:
        for finding in dc_report.findings:
            if not finding.get("vulnerable"):
                continue
            svc = finding.get("service", "")
            port = finding.get("port", "")
            findings.append({
                "id": f"dc-{dc_report.id}-{port}",
                "source": "default_creds",
                "category": "default_credentials",
                "severity": "Critical",
                "title": f"Default/no credentials on {svc} (port {port})",
                "description": (
                    f"Service {svc!r} on port {port} is accessible without authentication "
                    "or with vendor default credentials."
                ),
                "detail": {
                    "script": finding.get("script"),
                    "output_excerpt": (finding.get("output") or "")[:300],
                },
                "remediation": _default_creds_remediation(svc, port),
                "priority": 98,
            })

    # ── testssl findings ──────────────────────────────────────────────────────
    ts_result = await db.execute(
        select(TestsslReport)
        .where(TestsslReport.asset_id == asset_id, TestsslReport.status == "completed")
        .order_by(TestsslReport.created_at.desc())
        .limit(1)
    )
    ts_report = ts_result.scalar_one_or_none()
    if ts_report and ts_report.findings:
        for f in ts_report.findings:
            sev_raw = (f.get("severity") or "INFO").upper()
            sev = _testssl_sev_map(sev_raw)
            if sev == "Info":
                continue  # skip noise
            title = f.get("id", "tls-finding")
            findings.append({
                "id": f"ts-{ts_report.id}-{title}",
                "source": "testssl",
                "category": "tls",
                "severity": sev,
                "title": f"TLS: {title}",
                "description": f.get("finding") or title,
                "detail": {"port": ts_report.port, "grade": ts_report.grade},
                "remediation": _testssl_remediation(f),
                "priority": _severity_base_priority(sev),
            })

    # ── ssh-audit findings ────────────────────────────────────────────────────
    sa_result = await db.execute(
        select(SshAuditReport)
        .where(SshAuditReport.asset_id == asset_id, SshAuditReport.status == "completed")
        .order_by(SshAuditReport.created_at.desc())
        .limit(1)
    )
    sa_report = sa_result.scalar_one_or_none()
    if sa_report:
        for algo_list, category in (
            (sa_report.kex_algorithms, "kex"),
            (sa_report.encryption_algorithms, "encryption"),
            (sa_report.mac_algorithms, "mac"),
            (sa_report.host_key_algorithms, "host_key"),
        ):
            for algo in (algo_list or []):
                lvl = algo.get("level", "info")
                if lvl not in ("fail", "critical"):
                    continue
                sev = "Critical" if lvl == "critical" else "High"
                name = algo.get("algorithm", "unknown")
                findings.append({
                    "id": f"sa-{sa_report.id}-{category}-{name}",
                    "source": "ssh_audit",
                    "category": "ssh_config",
                    "severity": sev,
                    "title": f"SSH weak {category}: {name}",
                    "description": (
                        f"SSH server offers deprecated/weak {category} algorithm: {name!r}. "
                        "This may be exploitable by a network attacker."
                    ),
                    "detail": {"port": sa_report.port, "banner": sa_report.banner},
                    "remediation": _ssh_audit_remediation(category, name),
                    "priority": _severity_base_priority(sev),
                })
        # CVEs from ssh-audit
        raw = sa_report.raw_output or {}
        for cve_entry in (raw.get("cves") or []):
            cve_name = cve_entry.get("name", "unknown")
            findings.append({
                "id": f"sa-cve-{sa_report.id}-{cve_name}",
                "source": "ssh_audit",
                "category": "cve",
                "severity": "Critical",
                "title": cve_name,
                "description": cve_entry.get("description", ""),
                "detail": {"cvss": cve_entry.get("cvss"), "port": sa_report.port},
                "remediation": f"Upgrade SSH server to patched version. See {cve_name} advisory.",
                "priority": 92,
            })

    # ── Nuclei findings ───────────────────────────────────────────────────────
    nuc_result = await db.execute(
        select(NucleiReport)
        .where(NucleiReport.asset_id == asset_id, NucleiReport.status == "completed")
        .order_by(NucleiReport.created_at.desc())
        .limit(1)
    )
    nuc_report = nuc_result.scalar_one_or_none()
    if nuc_report and nuc_report.report:
        for f in (nuc_report.report.get("findings") or []):
            info = f.get("info", {})
            sev_raw = (info.get("severity") or "info").capitalize()
            if sev_raw.lower() in ("info", "unknown"):
                continue
            name = info.get("name") or f.get("template-id", "nuclei-finding")
            findings.append({
                "id": f"nuc-{nuc_report.id}-{f.get('template-id', '')}",
                "source": "nuclei",
                "category": "network_vuln",
                "severity": sev_raw,
                "title": name,
                "description": info.get("description") or name,
                "detail": {
                    "matched_at": f.get("matched-at"),
                    "tags": info.get("tags"),
                },
                "remediation": info.get("remediation") or _nuclei_remediation(f),
                "priority": _severity_base_priority(sev_raw),
            })

    # Deduplicate CVEs from Nuclei that are already in CVE findings
    cve_titles = {f["title"] for f in findings if f["category"] == "cve"}
    findings = [
        f for f in findings
        if not (f["source"] == "nuclei" and f["title"] in cve_titles)
    ]

    # Sort by priority descending
    findings.sort(key=lambda f: f["priority"], reverse=True)

    open_ports = [
        {"port": p.port_number, "protocol": p.protocol, "service": p.service_name, "version": p.version}
        for p in (asset.ports or []) if p.state == "open"
    ]

    return {
        "generated_at": now.isoformat(),
        "asset": {
            "id": str(asset.id),
            "name": asset.name,
            "ip": asset.ip,
            "hostname": asset.hostname,
            "os_family": asset.os_family,
            "os_version": asset.os_version,
            "device_type": asset.device_type,
            "criticality": asset.criticality,
            "risk_score": asset.risk_score,
            "tags": [t.name for t in (asset.tags or [])],
        },
        "open_ports": open_ports,
        "summary": {
            "total_findings": len(findings),
            "by_severity": _count_by_severity(findings),
            "by_category": _count_by_category(findings),
            "default_creds_found": sum(
                1 for f in findings if f["category"] == "default_credentials"
            ),
        },
        "findings": findings,
        "scan_coverage": {
            "testssl": ts_report.status if ts_report else "not_run",
            "ssh_audit": sa_report.status if sa_report else "not_run",
            "default_creds": dc_report.status if dc_report else "not_run",
            "nuclei": nuc_report.status if nuc_report else "not_run",
            "testssl_grade": ts_report.grade if ts_report else None,
        },
    }


# ── Priority scoring ──────────────────────────────────────────────────────────


def _severity_base_priority(severity: str) -> int:
    return {"Critical": 90, "High": 70, "Medium": 50, "Low": 30, "Info": 10}.get(
        severity.capitalize(), 10
    )


def _cve_priority(cve, link) -> int:
    base = _severity_base_priority(cve.severity or "Unknown")
    if link.exploit_verified:
        base += 8
    if cve.kev_date_added:
        base += 5
    if cve.exploit_maturity == "weaponized":
        base += 5
    elif cve.exploit_maturity == "exploit":
        base += 3
    if cve.epss_score and cve.epss_score > 0.5:
        base += 3
    if link.sla_breached:
        base += 4
    return min(base, 100)


# ── Remediation text generators ───────────────────────────────────────────────


def _cve_remediation(cve, link) -> str:
    parts = []
    if link.package_name and link.fixed_version:
        parts.append(
            f"Upgrade {link.package_name} from {link.package_version or '?'} "
            f"to {link.fixed_version} or later."
        )
    elif link.package_name:
        parts.append(f"Upgrade {link.package_name} ({link.package_version or 'installed'}) to latest.")
    if cve.remediation:
        parts.append(cve.remediation)
    if cve.kev_date_added:
        parts.append("Listed in CISA KEV — treat as actively exploited.")
    if not parts:
        parts.append(f"Review and patch {cve.cve_id}. Consult vendor advisory.")
    return " ".join(parts)


_DEFAULT_CREDS_REMED: dict[str, str] = {
    "ftp":        "Disable anonymous FTP. Require strong credentials. Consider SFTP instead.",
    "redis":      "Set 'requirepass <strong_password>' in redis.conf. Bind to 127.0.0.1. Use ACLs.",
    "mongodb":    "Enable MongoDB auth (--auth). Create admin user. Restrict bind IP.",
    "memcached":  "Bind memcached to 127.0.0.1 (-l 127.0.0.1). Firewall port 11211.",
    "mysql":      "Set a strong root password. Run 'mysql_secure_installation'. Remove anonymous users.",
    "mssql":      "Set a strong 'sa' password. Disable sa if possible. Use Windows Authentication.",
    "postgresql": "Set a strong postgres password. Configure pg_hba.conf to reject local trust.",
    "smtp":       "Disable SMTP open relay. Require authentication for relaying (SASL).",
    "snmp":       "Change default community strings ('public'/'private'). Upgrade to SNMPv3 with auth+priv.",
    "http":       "Change default admin credentials. Restrict admin interface by IP.",
    "https":      "Change default admin credentials. Restrict admin interface by IP.",
    "telnet":     "Disable Telnet. Replace with SSH.",
    "rdp":        "Enable Network Level Authentication (NLA). Use strong passwords + MFA.",
}


def _default_creds_remediation(service: str, port) -> str:
    for key in _DEFAULT_CREDS_REMED:
        if key in (service or "").lower():
            return _DEFAULT_CREDS_REMED[key]
    return f"Change default credentials for service on port {port}. Restrict network access."


def _testssl_remediation(finding: dict) -> str:
    fid = (finding.get("id") or "").lower()
    text = finding.get("finding") or ""
    if "heartbleed" in fid:
        return "Upgrade OpenSSL to patched version (>=1.0.1g). Regenerate private keys and certificates."
    if "poodle" in fid:
        return "Disable SSLv3. Enable TLS_FALLBACK_SCSV in your server configuration."
    if "beast" in fid:
        return "Prefer TLS 1.2+. If TLS 1.0 needed, prefer RC4 over CBC (or just disable TLS 1.0)."
    if "robot" in fid:
        return "Disable RSA key exchange in server cipher list. Use ECDHE/DHE ciphers only."
    if "tls1_0" in fid or "tls10" in fid:
        return "Disable TLS 1.0 and 1.1. Set minimum protocol to TLS 1.2 in server config."
    if "rc4" in fid or "export" in fid or "null" in fid:
        return "Remove weak/export/null ciphers from server cipher list. Use only AEAD ciphers."
    if "cert" in fid and ("expir" in text.lower() or "expired" in text.lower()):
        return "Renew the TLS certificate immediately. Configure automated renewal (Let's Encrypt/ACME)."
    if "hsts" in fid:
        return "Add 'Strict-Transport-Security: max-age=31536000; includeSubDomains' response header."
    return f"Reconfigure TLS to address: {text[:120] or fid}"


def _ssh_audit_remediation(category: str, algo: str) -> str:
    if "diffie-hellman-group1" in algo or "diffie-hellman-group14" in algo:
        return (
            f"Remove {algo!r} from KexAlgorithms in sshd_config. "
            "Use curve25519-sha256 or diffie-hellman-group16-sha512 instead."
        )
    if "arcfour" in algo or "blowfish" in algo or "3des" in algo:
        return (
            f"Remove {algo!r} from Ciphers in sshd_config. "
            "Use chacha20-poly1305@openssh.com or aes256-gcm@openssh.com."
        )
    if "hmac-md5" in algo or "hmac-sha1" in algo:
        return (
            f"Remove {algo!r} from MACs in sshd_config. "
            "Use hmac-sha2-256-etm@openssh.com or hmac-sha2-512-etm@openssh.com."
        )
    if "ssh-dss" in algo or "ssh-rsa" in algo:
        return (
            f"Remove {algo!r} from HostKeyAlgorithms. "
            "Use ed25519 or ecdsa-sha2-nistp521 host keys."
        )
    return f"Remove or replace SSH {category} algorithm {algo!r} as recommended by ssh-audit."


def _nuclei_remediation(finding: dict) -> str:
    info = finding.get("info", {})
    name = info.get("name") or finding.get("template-id", "this vulnerability")
    sev = (info.get("severity") or "").lower()
    refs = info.get("reference") or []
    ref_str = f" Reference: {refs[0]}" if refs else ""
    return f"Remediate {name!r} ({sev} severity).{ref_str} Apply vendor patch or configuration fix."


# ── Counting helpers ──────────────────────────────────────────────────────────


def _count_by_severity(findings: list[dict]) -> dict:
    counts: dict[str, int] = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Info": 0}
    for f in findings:
        sev = f.get("severity", "Info").capitalize()
        counts[sev] = counts.get(sev, 0) + 1
    return counts


def _count_by_category(findings: list[dict]) -> dict:
    counts: dict[str, int] = {}
    for f in findings:
        cat = f.get("category", "other")
        counts[cat] = counts.get(cat, 0) + 1
    return counts


def _testssl_sev_map(raw: str) -> str:
    return {
        "CRITICAL": "Critical",
        "HIGH": "High",
        "MEDIUM": "Medium",
        "WARN": "Medium",
        "LOW": "Low",
        "INFO": "Info",
        "OK": "Info",
        "DEBUG": "Info",
    }.get(raw, "Info")


# ── HTML rendering ────────────────────────────────────────────────────────────


def _render_asset_html(data: dict) -> str:
    asset = data["asset"]
    summary = data["summary"]
    findings = data["findings"]
    coverage = data["scan_coverage"]

    sev_colors = {
        "Critical": ("#ffeaea", "#c0392b"),
        "High":     ("#fff3e0", "#e67e22"),
        "Medium":   ("#fffde7", "#f39c12"),
        "Low":      ("#e8f5e9", "#27ae60"),
        "Info":     ("#f0f4ff", "#3498db"),
    }

    risk_score = asset.get("risk_score")
    risk_color = (
        "#c0392b" if risk_score and risk_score >= 70
        else "#e67e22" if risk_score and risk_score >= 40
        else "#27ae60"
    )

    def badge(sev: str) -> str:
        bg, fg = sev_colors.get(sev, ("#eee", "#333"))
        return (
            f'<span style="background:{bg};color:{fg};padding:2px 8px;border-radius:4px;'
            f'font-size:.75rem;font-weight:700;text-transform:uppercase">{sev}</span>'
        )

    def coverage_badge(status: str) -> str:
        if status == "completed":
            return '<span style="color:#27ae60;font-weight:600">✓ done</span>'
        if status == "not_run":
            return '<span style="color:#aaa">— not run</span>'
        return f'<span style="color:#e67e22">{status}</span>'

    findings_html = ""
    for i, f in enumerate(findings):
        bg, fg = sev_colors.get(f["severity"], ("#eee", "#333"))
        detail_items = "".join(
            f"<tr><td style='color:#888;padding:3px 8px;white-space:nowrap'>{k}</td>"
            f"<td style='padding:3px 8px'>{v}</td></tr>"
            for k, v in (f.get("detail") or {}).items()
            if v is not None and v != "" and v is not False
        )
        detail_block = (
            f'<table style="font-size:.8rem;margin-top:8px;border-collapse:collapse">'
            f'{detail_items}</table>'
            if detail_items else ""
        )
        findings_html += f"""
<div style="border-left:4px solid {fg};background:{bg};border-radius:6px;
            padding:14px 18px;margin-bottom:12px;">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
    {badge(f['severity'])}
    <span style="font-size:.75rem;color:#888">[{f['source']}]</span>
    <strong style="font-size:.95rem">{f['title']}</strong>
    <span style="margin-left:auto;font-size:.75rem;color:#aaa">prio {f['priority']}</span>
  </div>
  <p style="font-size:.88rem;color:#444;margin-bottom:6px">{f['description']}</p>
  {detail_block}
  <div style="margin-top:10px;background:rgba(255,255,255,.6);border-radius:4px;
              padding:8px 12px;font-size:.83rem">
    <strong style="color:#1a2332">Remédiation :</strong> {f['remediation']}
  </div>
</div>"""

    port_rows = "".join(
        f"<tr><td>{p['port']}/{p['protocol']}</td><td>{p['service'] or '—'}</td>"
        f"<td style='color:#888'>{p['version'] or '—'}</td></tr>"
        for p in data["open_ports"]
    )

    sev_summary = "".join(
        f'<div style="text-align:center;padding:10px 16px"><div style="font-size:1.8rem;'
        f'font-weight:700;color:{sev_colors.get(s, ("","#333"))[1]}">{summary["by_severity"].get(s,0)}</div>'
        f'<div style="font-size:.75rem;color:#888">{s}</div></div>'
        for s in ("Critical", "High", "Medium", "Low")
    )

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<title>Audit — {asset.get('name') or asset.get('ip')}</title>
<style>
  * {{ box-sizing:border-box;margin:0;padding:0 }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
          background:#f4f6f9;color:#333 }}
  header {{ background:#1a2332;color:#fff;padding:24px 40px }}
  header h1 {{ font-size:1.6rem }}
  header p {{ opacity:.7;font-size:.85rem;margin-top:4px }}
  main {{ max-width:1000px;margin:28px auto;padding:0 20px }}
  .card {{ background:#fff;border-radius:8px;padding:20px 24px;
           box-shadow:0 1px 4px rgba(0,0,0,.1);margin-bottom:20px }}
  .card h2 {{ font-size:1rem;color:#1a2332;margin-bottom:14px;
              border-bottom:1px solid #eee;padding-bottom:8px }}
  table {{ width:100%;border-collapse:collapse;font-size:.88rem }}
  th {{ background:#f0f2f5;padding:7px 10px;text-align:left;font-weight:600 }}
  td {{ padding:6px 10px;border-bottom:1px solid #f0f0f0 }}
  tr:last-child td {{ border-bottom:none }}
  .meta-grid {{ display:grid;grid-template-columns:repeat(3,1fr);gap:12px }}
  .meta-item label {{ font-size:.75rem;color:#888;display:block }}
  .meta-item span {{ font-size:.95rem;font-weight:600 }}
  footer {{ text-align:center;color:#999;font-size:.78rem;padding:20px }}
</style>
</head>
<body>
<header>
  <h1>Rapport d'audit de sécurité — {asset.get('name') or asset.get('ip') or 'Asset'}</h1>
  <p>Généré le {data['generated_at'][:19].replace('T',' ')} UTC</p>
</header>
<main>

<div class="card">
  <h2>Informations sur l'asset</h2>
  <div class="meta-grid">
    <div class="meta-item"><label>IP</label><span>{asset.get('ip') or '—'}</span></div>
    <div class="meta-item"><label>Hostname</label><span>{asset.get('hostname') or '—'}</span></div>
    <div class="meta-item"><label>OS</label><span>{asset.get('os_family') or '—'} {asset.get('os_version') or ''}</span></div>
    <div class="meta-item"><label>Criticité</label><span>{(asset.get('criticality') or '').capitalize()}</span></div>
    <div class="meta-item"><label>Risk Score</label>
      <span style="color:{risk_color};font-size:1.3rem">{f"{risk_score:.1f}/100" if risk_score is not None else "—"}</span>
    </div>
    <div class="meta-item"><label>Tags</label><span>{', '.join(asset.get('tags') or []) or '—'}</span></div>
  </div>
</div>

<div class="card">
  <h2>Couverture des scans</h2>
  <table>
    <thead><tr><th>Scanner</th><th>Statut</th><th>Info</th></tr></thead>
    <tbody>
      <tr><td>testssl.sh</td><td>{coverage_badge(coverage['testssl'])}</td>
          <td>{f"Grade: <strong>{coverage['testssl_grade']}</strong>" if coverage.get('testssl_grade') else '—'}</td></tr>
      <tr><td>ssh-audit</td><td>{coverage_badge(coverage['ssh_audit'])}</td><td>—</td></tr>
      <tr><td>Default credentials</td><td>{coverage_badge(coverage['default_creds'])}</td><td>—</td></tr>
      <tr><td>Nuclei</td><td>{coverage_badge(coverage['nuclei'])}</td><td>—</td></tr>
    </tbody>
  </table>
</div>

<div class="card">
  <h2>Résumé des findings</h2>
  <div style="display:flex;gap:0;border-radius:6px;overflow:hidden;border:1px solid #eee">
    {sev_summary}
  </div>
</div>

<div class="card">
  <h2>Ports ouverts ({len(data['open_ports'])})</h2>
  {'<table><thead><tr><th>Port</th><th>Service</th><th>Version</th></tr></thead><tbody>' + port_rows + '</tbody></table>' if data['open_ports'] else '<p style="color:#888;font-size:.88rem">Aucun port scanné</p>'}
</div>

<div class="card">
  <h2>Findings par ordre de priorité ({summary['total_findings']})</h2>
  {findings_html if findings_html else '<p style="color:#27ae60;font-weight:600">Aucun finding — asset propre ✓</p>'}
</div>

</main>
<footer>NetLanVentory — Confidentiel — {date.today().isoformat()}</footer>
</body>
</html>"""
